from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def gerar_relatorio_excel(
    estatisticas,
    identificador_execucao: str,
    inicio_execucao,
    fim_execucao
):
    pasta = Path("relatorios")
    pasta.mkdir(parents=True, exist_ok=True)

    caminho = pasta / (
        f"relatorio_{identificador_execucao}.xlsx"
    )

    workbook = Workbook()

    aba_resultados = workbook.active
    aba_resultados.title = "Resultados"

    cabecalhos = [
        "Centro Operativo",
        "Código Externo",
        "Resultado",
        "Mensagem"
    ]

    aba_resultados.append(cabecalhos)

    for resultado in estatisticas.resultados:
        aba_resultados.append([
            resultado["centro_operativo"],
            resultado["codigo_externo"],
            resultado["resultado"],
            resultado["mensagem"]
        ])

    preenchimento_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    fonte_cabecalho = Font(
        color="FFFFFF",
        bold=True
    )

    for celula in aba_resultados[1]:
        celula.fill = preenchimento_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(
            horizontal="center"
        )

    larguras = [20, 24, 20, 70]

    for indice, largura in enumerate(larguras, start=1):
        coluna = get_column_letter(indice)
        aba_resultados.column_dimensions[coluna].width = largura

    aba_resultados.freeze_panes = "A2"
    aba_resultados.auto_filter.ref = aba_resultados.dimensions

    aba_resumo = workbook.create_sheet("Resumo")

    tempo_total = fim_execucao - inicio_execucao

    dados_resumo = [
        ["Execução", identificador_execucao],
        ["Início", inicio_execucao.strftime("%d/%m/%Y %H:%M:%S")],
        ["Fim", fim_execucao.strftime("%d/%m/%Y %H:%M:%S")],
        ["Tempo total", str(tempo_total)],
        ["Total de TdCs", estatisticas.total],
        ["Processadas", estatisticas.processadas],
        ["Canceladas", estatisticas.canceladas],
        ["Já anuladas", estatisticas.ja_anuladas],
        ["WO finalizadas", estatisticas.wo_finalizadas],
        ["Erros", estatisticas.erros]
    ]

    for linha in dados_resumo:
        aba_resumo.append(linha)

    aba_resumo.column_dimensions["A"].width = 25
    aba_resumo.column_dimensions["B"].width = 30

    for celula in aba_resumo["A"]:
        celula.font = Font(bold=True)

    workbook.save(caminho)

    return caminho