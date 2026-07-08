from openpyxl import load_workbook


def ler_notas_excel(caminho_arquivo: str):
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook.active

    notas = []

    for linha in range(2, sheet.max_row + 1):
        centro_operativo = sheet[f"A{linha}"].value
        codigo_externo = sheet[f"E{linha}"].value

        if not centro_operativo or not codigo_externo:
            continue

        notas.append({
            "linha": linha,
            "centro_operativo": str(centro_operativo).strip(),
            "codigo_externo": str(codigo_externo).strip()
        })

    return notas